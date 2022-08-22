import random
from openpyxl import load_workbook 
wb=load_workbook('trynewdata.xlsx')
ws1=wb.active
sno=ws1.max_row-1
details=[]
totalbill=[]
print('WELCOME TO HOTEL THE PARADISE INN','\n')
def mainmenu():
    print('MAIN MENU')
    print("1.BOOK A ROOM")
    print("2.CALCULATE ROOMRENT")
    print("3.RESTAURANT BILL")
    print("4.GAME BILL")
    print('5.SHOW MY BOOKINGS AND BILLS')
    print("6.SHOW TOTAL COST")
    print("7.EXIT")
    ch=int(input("ENTER YOUR CHOICE:"))
    if ch==1:
        name=input("Enter your name:")
        address=input("Enter your address:")
        checkin=input("Enter your check-in date:")
        checkout=input("Enter your check-out date:")
        roomno=random.randint(100,5000)
        print('Congratulations, your room is booked!')
        print('Your room number is:',roomno)
        details.extend([sno,name,address,checkin,checkout,roomno])
        mainmenu()
    if ch==2:
        print('\n',"Rooms are as follows:")
        print("1.Type A = Rs.6000 per night","2.Type B = Rs.5000 per night","3.Type C = Rs.4000 per night","4.Type D = Rs.3000 per night",'\n',sep="\n")
        r=int(input("Enter room type:"))
        night=int(input("No of nights you stayed:"))
        if r==1:
            r1=6000*night
            r2,r3,r4=0,0,0
            print("Your room rent is Rs.",r1)
            Rtype='Type A-> Rs.6000 per night'
            details.extend([Rtype,night])
        elif r==2:
            r2=5000*night
            r1,r3,r4=0,0,0
            print("Your room rent is Rs.",r2)
            Rtype='Type B-> Rs.5000 per night'
            details.extend([Rtype,night])
        elif r==3:
            r3=4000*night
            r2,r1,r4=0,0,0
            print("Your room rent is Rs.",r3)
            Rtype='Type C-> Rs.4000 per night'
            details.extend([Rtype,night])
        elif r==4:
            r4=3000*night
            r2,r3,r1=0,0,0
            print("Your room rent is Rs.",r4)
            Rtype='Type D-> Rs.3000 per night'
            details.extend([Rtype,night])
        else:
            print('Invalid choice. Please enter between 1-4')
            mainmenu()
        RRENT=r1+r2+r3+r4
        details.append(RRENT)
        totalbill.append(RRENT)
        mainmenu()
    if ch==3:
        print('FOOD MENU:')
        print("1.Water->Rs.20","2.Tea->Rs.10","3.Lunch->Rs.110","4.Dinner->Rs.150",sep="\n")
        n1=int(input("Enter no. of items you want:"))
        foodbill=0
        food=()
        food1=''
        for i in range(1,n1+1):
            rb=int(input("Enter Your Choice:"))
            if rb==1:
                q=int(input("Enter the quantity:"))
                foodbill=foodbill+q*20
                dish='Water'
                price=20
            elif rb==2:
                q=int(input("Enter the quantity:"))
                foodbill=foodbill+q*10
                dish='Tea'
                price=10
            elif rb==3:
                q=int(input("Enter the quantity:"))
                foodbill=foodbill+q*110
                dish='Lunch'
                price=110
            elif rb==4:
                q=int(input("Enter the quantity:"))
                foodbill=foodbill+q*150
                dish='Dinner'
                price=150
            else:
                print('Invalid choice. Please enter between 1-4')
                mainmenu()
            res=(dish+' ['+'Quantity:'+str(q)+','+'Rate:'+'Rs.'+str(price)+']',)
            food=food+res
        print('Total restaurant bill is Rs.',foodbill)
        a=str(food)
        for i in a:
            if i=='(' or i==')' or i=="'":
                food1=food1+''
            else:
                food1=food1+i
        details.extend([n1,food1,foodbill])
        totalbill.append(foodbill)
        mainmenu()
    if ch==4:
        print('\n','Numerous games available are:')
        print("1.Bowling->Rs800","2.Smash->Rs.1000","3.Call Of Duty->Rs.200","4.God Of War->Rs.200","5.Adventure Park->Rs.1500",sep="\n")
        n2=int(input("Enter no. of games you want:"))
        gamebill=0
        gt=()
        gt1=''
        for j in range(1,n2+1):
            y=int(input("Enter Your Choice:"))
            if y==1:
                gamebill=gamebill+800
                game='Bowling'
                rate=800
            elif y==2:
                gamebill=gamebill+1000
                game='Smash'
                rate=1000
            elif y==3:
                gamebill=gamebill+200
                game='Call Of Duty'
                rate=200
            elif y==4:
                gamebill=gamebill+200
                game='God Of War'
                rate=200
            elif y==5:
                gamebill=gamebill+1500
                game='Adventure Park'
                rate=1500
            else:
                print('Invalid choice. Please enter between 1-5')
                mainmenu()
            g=(game+' ['+'Rate: Rs.'+str(rate)+']',)
            gt=gt+g
        print("Total game bill is Rs.",gamebill)
        x=str(gt)
        for j in x:
            if j=='(' or j==')'or j=="'":
                gt1=gt1+''
            else:
                gt1=gt1+j
        details.extend([n2,gt1,gamebill])
        totalbill.append(gamebill)
        mainmenu() 
    #if ch==5:
        #show()
    #if ch==6:
        #tcost()
    if ch==7:
        sumbill=0
        for k in totalbill:
            sumbill=sumbill+k
        details.append(sumbill)
        ws1.append(details)
        wb.save(filename='trynewdata.xlsx')
        exit()
    else:
        print('Invalid choice. Please enter between 1-7')
        mainmenu()
mainmenu()




    

    
  

    
    
    









